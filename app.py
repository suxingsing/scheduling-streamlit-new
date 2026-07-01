import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import io
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ------------------------------
# 一、全局核心配置
# ------------------------------
# 制程配置
PROCESS_CONFIG = {
    "SMT": {"default_uph": 1200, "default_lead_time": 3, "default_work_hours": 12, "desc": "贴片制程"},
    "预组": {"default_uph": 600, "default_lead_time": 2, "default_work_hours": 10, "desc": "预组装制程"},
    "组装": {"default_uph": 400, "default_lead_time": 2, "default_work_hours": 10, "desc": "整机组装制程"},
    "二测": {"default_uph": 800, "default_lead_time": 1, "default_work_hours": 10, "desc": "二次测试制程"},
    "包装": {"default_uph": 1500, "default_lead_time": 1, "default_work_hours": 10, "desc": "成品包装制程"}
}

# 爬坡斜率数据（DAY1到DAYn，产能从低到高递增，倒排时严格从DAY1开始推算）
RAMP_DATA = {
    "高端机": {
        "100%新人": [10, 20, 30, 40, 50, 60, 65, 70, 75, 80, 85, 90, 95, 95, 100],
        "50%新人": [15, 25, 35, 45, 55, 65, 75, 80, 85, 90, 95, 100],
        "0%新人": [20, 30, 40, 50, 60, 70, 80, 90, 95, 100]
    },
    "中端机": {
        "100%新人": [10, 20, 30, 40, 50, 60, 70, 80, 85, 90, 95, 100],
        "50%新人": [15, 30, 40, 50, 60, 70, 80, 90, 100],
        "0%新人": [20, 40, 60, 70, 80, 90, 100]
    },
    "低端机": {
        "100%新人": [20, 30, 40, 50, 60, 70, 80, 90, 95, 100],
        "50%新人": [30, 50, 70, 80, 85, 90, 95, 100],
        "0%新人": [40, 60, 80, 90, 95, 100]
    }
}

# ------------------------------
# 二、核心工具函数
# ------------------------------
def generate_full_date_list(start_date, end_date):
    """生成完整的自然日列表"""
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    return date_list

def get_previous_workday(base_date, days_to_go_back, rest_dates_set):
    """从基准日期往前数N个工作日，返回目标日期"""
    current_date = base_date
    count = 0
    while count < days_to_go_back:
        current_date -= timedelta(days=1)
        if current_date not in rest_dates_set:
            count += 1
    return current_date

def get_next_workday(base_date, days_to_go_forward, rest_dates_set):
    """从基准日期往后数N个工作日，返回目标日期"""
    current_date = base_date
    count = 0
    while count < days_to_go_forward:
        current_date += timedelta(days=1)
        if current_date not in rest_dates_set:
            count += 1
    return current_date

def short_date_label(d):
    return f"{d.month}月{d.day}日"

def parse_template_date(value, fallback_year):
    """兼容模板里的 4月1日、4/1、2026-04-01 等日期写法。"""
    if pd.isna(value):
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            return pd.to_datetime(value, unit="D", origin="1899-12-30").date()
        except Exception:
            return None

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            year = parsed.year if "%Y" in fmt else fallback_year
            return parsed.replace(year=year).date()
        except ValueError:
            pass
    if "月" in text and "日" in text:
        try:
            month = int(text.split("月", 1)[0])
            day = int(text.split("月", 1)[1].split("日", 1)[0])
            return datetime(fallback_year, month, day).date()
        except Exception:
            return None
    return None

def build_material_template(start_date, end_date, default_initial_stock):
    """生成横排日期物料交期模板，日期范围跟随页面输入。"""
    date_list = generate_full_date_list(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "物料交期输入"

    ws["A1"] = "物料交期"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(date_list) + 1))
    ws["A3"] = "物料期初库存"
    ws["A4"] = int(default_initial_stock)

    header_fill = PatternFill("solid", fgColor="C9B8AA")
    input_fill = PatternFill("solid", fgColor="F2DED2")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_idx, d in enumerate(date_list, start=2):
        ws.cell(row=3, column=col_idx, value=short_date_label(d))
        ws.cell(row=4, column=col_idx, value=0)

    for row_idx in range(3, 5):
        for col_idx in range(1, len(date_list) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            elif col_idx == 1:
                cell.fill = input_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = value_fill

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 13

    help_ws = wb.create_sheet("填写说明")
    help_ws.append(["字段", "填写要求"])
    help_ws.append(["物料期初库存", "填写排程开始前可用物料数量，位于 A4。"])
    help_ws.append(["日期", "第 3 行由系统按页面排产开始日期和需求最终截止日期横向生成。"])
    help_ws.append(["预计到料数量", "第 4 行填写每天预计到料数量，空白按 0 处理。"])
    help_ws.append(["到料规则", "T 日到料最早 T+1 日投入排产。"])
    help_ws.column_dimensions["A"].width = 18
    help_ws.column_dimensions["B"].width = 58

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def parse_material_upload(uploaded_file, start_date, end_date, default_initial_stock):
    """读取竖排模板；同时兼容旧横排模板和普通两列表。"""
    fallback_year = start_date.year
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    header_row_idx = None
    for idx, row in raw_df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if isinstance(first_cell, str) and "物料期初库存" in first_cell:
            header_row_idx = idx
            break

    if header_row_idx is not None and header_row_idx + 1 < len(raw_df):
        first_value = raw_df.iloc[header_row_idx, 1] if raw_df.shape[1] > 1 else None
        next_first = raw_df.iloc[header_row_idx + 1, 0]
        date_header_idx = header_row_idx + 1
        if isinstance(next_first, str) and "半成品" in next_first:
            date_header_idx = header_row_idx + 2
            next_first = raw_df.iloc[date_header_idx, 0] if date_header_idx < len(raw_df) else None

        if isinstance(next_first, str) and "日期" in next_first:
            material_initial_stock = default_initial_stock if pd.isna(first_value) else int(first_value)
            records = []
            invalid_cells = 0
            for row_idx in range(date_header_idx + 1, len(raw_df)):
                d = parse_template_date(raw_df.iloc[row_idx, 0], fallback_year)
                if d is None:
                    continue
                qty_value = raw_df.iloc[row_idx, 1] if raw_df.shape[1] > 1 else 0
                if pd.isna(qty_value) or qty_value == "":
                    qty = 0
                else:
                    try:
                        qty = int(qty_value)
                        if qty < 0:
                            raise ValueError
                    except Exception:
                        invalid_cells += 1
                        qty = 0
                records.append({"日期": d, "预计到料数量": qty})
            messages = []
            if invalid_cells:
                messages.append(f"有 {invalid_cells} 个到料数量无法识别，已按 0 处理。")
            return material_initial_stock, pd.DataFrame(records), messages

        date_row = raw_df.iloc[header_row_idx]
        value_row = raw_df.iloc[header_row_idx + 1]
        raw_initial_stock = value_row.iloc[0]
        material_initial_stock = default_initial_stock if pd.isna(raw_initial_stock) else int(raw_initial_stock)
        records = []
        invalid_cells = 0
        for col_idx in range(1, len(date_row)):
            d = parse_template_date(date_row.iloc[col_idx], fallback_year)
            if d is None:
                continue
            qty_value = value_row.iloc[col_idx] if col_idx < len(value_row) else 0
            if pd.isna(qty_value) or qty_value == "":
                qty = 0
            else:
                try:
                    qty = int(qty_value)
                    if qty < 0:
                        raise ValueError
                except Exception:
                    invalid_cells += 1
                    qty = 0
            records.append({"日期": d, "预计到料数量": qty})
        messages = []
        if invalid_cells:
            messages.append(f"有 {invalid_cells} 个到料数量无法识别，已按 0 处理。")
        return material_initial_stock, pd.DataFrame(records), messages

    uploaded_file.seek(0)
    table_df = pd.read_excel(uploaded_file)
    if {"日期", "预计到料数量"}.issubset(table_df.columns):
        material_initial_stock = default_initial_stock
        if "物料期初库存" in table_df.columns and not table_df["物料期初库存"].dropna().empty:
            material_initial_stock = int(table_df["物料期初库存"].dropna().iloc[0])
        material_df = table_df[["日期", "预计到料数量"]].copy()
        material_df["日期"] = pd.to_datetime(material_df["日期"]).dt.date
        material_df["预计到料数量"] = material_df["预计到料数量"].fillna(0).astype(int)
        return material_initial_stock, material_df, []

    raise ValueError("未找到“物料期初库存 + 日期列”的模板结构，也未找到“日期/预计到料数量”两列。")

def parse_material_plan(material_plan_df):
    plan = {}
    invalid_rows = 0
    if material_plan_df is None:
        return plan, invalid_rows
    for _, row in material_plan_df.iterrows():
        raw_date = row.get("日期")
        raw_qty = row.get("预计到料数量", row.get("物料交期"))
        if pd.isna(raw_date) and pd.isna(raw_qty):
            continue
        try:
            d = raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
            qty = 0 if pd.isna(raw_qty) else int(raw_qty)
        except Exception:
            invalid_rows += 1
            continue
        if qty < 0:
            invalid_rows += 1
            continue
        plan[d] = plan.get(d, 0) + qty
    return plan, invalid_rows

def build_work_hours_template(start_date, end_date, default_work_hours):
    """生成横排日期单日工时模板，日期范围跟随页面输入。"""
    date_list = generate_full_date_list(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "单日工时输入"

    ws["A1"] = "单日工时"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(date_list) + 1))
    ws["A3"] = "日期"
    ws["A4"] = "单班组单日工时"

    header_fill = PatternFill("solid", fgColor="C9B8AA")
    input_fill = PatternFill("solid", fgColor="F2DED2")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_idx, d in enumerate(date_list, start=2):
        ws.cell(row=3, column=col_idx, value=short_date_label(d))
        ws.cell(row=4, column=col_idx, value=int(default_work_hours))

    for row_idx in range(3, 5):
        for col_idx in range(1, len(date_list) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            elif col_idx == 1:
                cell.fill = input_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = value_fill

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 12

    help_ws = wb.create_sheet("填写说明")
    help_ws.append(["字段", "填写要求"])
    help_ws.append(["日期", "由系统按页面排产开始日期和需求最终截止日期横向自动生成。"])
    help_ws.append(["单班组单日工时", "在每个日期下方填写当天单班组工作小时，UPH保持不变，空白沿用默认工时。"])
    help_ws.column_dimensions["A"].width = 18
    help_ws.column_dimensions["B"].width = 58

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def parse_work_hours_upload(uploaded_file, start_date, end_date, default_work_hours):
    fallback_year = start_date.year
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    horizontal_date_row_idx = None
    for idx, row in raw_df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        if isinstance(first_cell, str) and "日期" in first_cell:
            parsed_dates = [
                parse_template_date(value, fallback_year)
                for value in row.iloc[1:].tolist()
            ]
            if any(d is not None for d in parsed_dates):
                horizontal_date_row_idx = idx
                break

    if horizontal_date_row_idx is not None:
        hour_row_idx = horizontal_date_row_idx + 1
        if hour_row_idx >= len(raw_df):
            raise ValueError("工时模板缺少“单班组单日工时”填写行。")
        records = []
        invalid_cells = 0
        for col_idx in range(1, raw_df.shape[1]):
            d = parse_template_date(raw_df.iloc[horizontal_date_row_idx, col_idx], fallback_year)
            if d is None:
                continue
            hour_value = raw_df.iloc[hour_row_idx, col_idx]
            if pd.isna(hour_value) or hour_value == "":
                hours = int(default_work_hours)
            else:
                try:
                    hours = int(hour_value)
                    if hours < 0 or hours > 24:
                        raise ValueError
                except Exception:
                    invalid_cells += 1
                    hours = int(default_work_hours)
            records.append({"日期": d, "单班组单日工时": hours})
        messages = []
        if invalid_cells:
            messages.append(f"有 {invalid_cells} 个工时无法识别，已沿用默认工时。")
        return pd.DataFrame(records), messages

    header_row_idx = None
    for idx, row in raw_df.iterrows():
        first_cell = row.iloc[0] if len(row) else None
        second_cell = row.iloc[1] if len(row) > 1 else None
        if isinstance(first_cell, str) and "日期" in first_cell and isinstance(second_cell, str) and "工时" in second_cell:
            header_row_idx = idx
            break

    if header_row_idx is not None:
        records = []
        invalid_cells = 0
        for row_idx in range(header_row_idx + 1, len(raw_df)):
            d = parse_template_date(raw_df.iloc[row_idx, 0], fallback_year)
            if d is None:
                continue
            hour_value = raw_df.iloc[row_idx, 1] if raw_df.shape[1] > 1 else default_work_hours
            if pd.isna(hour_value) or hour_value == "":
                hours = int(default_work_hours)
            else:
                try:
                    hours = int(hour_value)
                    if hours < 0 or hours > 24:
                        raise ValueError
                except Exception:
                    invalid_cells += 1
                    hours = int(default_work_hours)
            records.append({"日期": d, "单班组单日工时": hours})
        messages = []
        if invalid_cells:
            messages.append(f"有 {invalid_cells} 个工时无法识别，已沿用默认工时。")
        return pd.DataFrame(records), messages

    uploaded_file.seek(0)
    table_df = pd.read_excel(uploaded_file)
    if {"日期", "单班组单日工时"}.issubset(table_df.columns):
        hours_df = table_df[["日期", "单班组单日工时"]].copy()
        hours_df["日期"] = pd.to_datetime(hours_df["日期"]).dt.date
        hours_df["单班组单日工时"] = hours_df["单班组单日工时"].fillna(default_work_hours).astype(int).clip(0, 24)
        return hours_df, []

    raise ValueError("未找到“日期 / 单班组单日工时”两列。")

def parse_work_hours_plan(work_hours_plan_df, default_work_hours):
    plan = {}
    invalid_rows = 0
    if work_hours_plan_df is None:
        return plan, invalid_rows
    for _, row in work_hours_plan_df.iterrows():
        raw_date = row.get("日期")
        raw_hours = row.get("单班组单日工时", row.get("工时"))
        if pd.isna(raw_date) and pd.isna(raw_hours):
            continue
        try:
            d = raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
            hours = default_work_hours if pd.isna(raw_hours) else int(raw_hours)
            if hours < 0 or hours > 24:
                raise ValueError
        except Exception:
            invalid_rows += 1
            continue
        plan[d] = hours
    return plan, invalid_rows

def is_shift_empty(shift_daily_prod):
    """判断一个班组是否全月为空，用于隐藏空班组"""
    for val in shift_daily_prod:
        if isinstance(val, (int, float)) and val > 0:
            return False
    return True

def calculate_shift_total_production(shift_daily_prod):
    """计算一个班组的总产量，用于优化分析"""
    total = 0
    for val in shift_daily_prod:
        if isinstance(val, (int, float)):
            total += val
    return total

def to_int_or_zero(value):
    try:
        if pd.isna(value):
            return 0
        numeric_value = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric_value):
            return 0
        return int(numeric_value)
    except Exception:
        return 0

def build_schedule_insights(
    schedule_df,
    production_target,
    total_demand,
    initial_stock,
    special_occupy,
    total_exist_capacity,
    existing_shift_count,
    uph_base,
    material_schedule_enabled,
    material_warnings,
    mode,
    idle_convert_threshold_days=5,
):
    """基于排产结果表生成结论和建议，避免只展示明细表。"""
    if schedule_df.empty:
        return {}, [], []

    date_cols = list(schedule_df.columns[2:])
    shift_rows = schedule_df[schedule_df["班组/指标"].astype(str).str.startswith("班组", na=False)]
    old_shift_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("老班组", na=False)]
    new_shift_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("新班组", na=False)]

    daily_totals = []
    for col in date_cols:
        daily_totals.append(sum(to_int_or_zero(row[col]) for _, row in shift_rows.iterrows()))
    planned_output = sum(daily_totals)
    shortage = max(int(production_target) - planned_output, 0)
    original_required_output = max(int(total_demand) + int(special_occupy) - int(initial_stock), 0)
    material_shortage_qty = max(original_required_output - int(production_target), 0) if material_schedule_enabled else 0

    last_prod_day = "无"
    for col, qty in zip(date_cols, daily_totals):
        if qty > 0:
            last_prod_day = col.replace("\n", " ")

    hours_rows = schedule_df[schedule_df["班组/指标"].astype(str) == "单班组单日工时"]
    if not hours_rows.empty:
        hours_row = hours_rows.iloc[0]
        daily_capacity_map = {col: to_int_or_zero(hours_row[col]) * int(uph_base) for col in date_cols}
    else:
        daily_capacity_map = {col: 0 for col in date_cols}

    old_idle_days = 0
    for _, row in old_shift_rows.iterrows():
        for col in date_cols:
            day_capacity = daily_capacity_map.get(col, 0)
            # 放空只统计非休息日且有排产工时的日期；休息日或 0 工时日期不计入放空。
            if day_capacity <= 0:
                continue
            val = row[col]
            if str(val).strip() == "":
                continue
            produced_qty = to_int_or_zero(val)
            if str(val) == "当日放空" or produced_qty < day_capacity:
                old_idle_days += 1

    material_gap_min = None
    if material_schedule_enabled and "累计物料gap" in schedule_df["班组/指标"].astype(str).values:
        gap_row = schedule_df[schedule_df["班组/指标"].astype(str) == "累计物料gap"].iloc[0]
        material_gap_values = [to_int_or_zero(gap_row[col]) for col in date_cols]
        material_gap_min = min(material_gap_values) if material_gap_values else None

    summary = {
        "planned_output": planned_output,
        "shortage": shortage,
        "old_shift_count": len(old_shift_rows),
        "new_shift_count": len(new_shift_rows),
        "last_prod_day": last_prod_day,
        "old_idle_days": old_idle_days,
        "material_gap_min": material_gap_min,
        "material_shortage_qty": material_shortage_qty,
    }

    conclusions = []
    if shortage > 0:
        conclusions.append(f"本次排产未完全覆盖目标，还差 {shortage:,} 件。")
    else:
        conclusions.append(f"本次排产可覆盖目标，计划产出 {planned_output:,} 件，目标 {int(production_target):,} 件。")

    if len(new_shift_rows) > 0:
        conclusions.append(f"系统启用了 {len(new_shift_rows)} 个新班组，新班组按爬坡曲线从启用日连续正排。")
    else:
        conclusions.append(f"当前方案仅使用老班组，启用老班组 {len(old_shift_rows)} 个。")

    if old_idle_days > 0:
        conclusions.append(f"老班组存在 {old_idle_days} 个工作日放空/未排满日期；统计时已排除休息日及 0 工时日期。")

    if material_schedule_enabled:
        conclusions.append("本次已按上传的物料交期约束排产，物料到料按 T+1 可用处理。")
    else:
        conclusions.append("本次未启用物料交期约束，排产结果仅按产能、工时、日历与爬坡规则计算。")

    suggestions = []
    if shortage > 0:
        suggestions.append("优先补充物料到料、增加可用工时或延长排产周期，然后重新排产。")
    elif "新班组" in mode:
        suggestions.append("建议复核新增班组的人员到位时间和爬坡比例，确认连续正排启动日期可执行。")
    elif total_exist_capacity > production_target:
        suggestions.append("当前目标已覆盖，可按业务需要复核班组启用数量或加班安排。")

    if old_idle_days > idle_convert_threshold_days:
        suggestions.append(f"老班组工作日放空天数超过 {idle_convert_threshold_days} 天，建议复核需求节奏、物料交期或减班策略。")
    elif old_idle_days > 0:
        suggestions.append("老班组存在少量工作日放空，可结合现场换线、人员调配或日工时调整进一步平滑产能。")

    if material_schedule_enabled and material_gap_min is not None and material_gap_min <= 0:
        suggestions.append("物料 gap 已接近 0，建议复核关键日期前一日到料，避免实际到料延迟导致排产中断。")
    if not material_schedule_enabled:
        suggestions.append("正式使用前建议上传物料交期 Excel，避免产能计划与真实到料节奏脱节。")

    for warning in material_warnings:
        if warning not in suggestions:
            suggestions.append(warning)

    return summary, conclusions, suggestions

def calculate_ramp_need_days(final_gap, ramp_curve, single_shift_daily):
    """
    核心倒排函数：根据缺口计算需要的爬坡天数和产能序列
    规则：从DAY1开始累加爬坡产能，直到覆盖缺口，返回需要的天数和产能列表
    """
    if final_gap <= 0:
        return 0, []

    cumulative_prod = 0
    need_days = 0
    prod_list = []

    # 从DAY1开始累加，直到覆盖缺口
    for rate in ramp_curve:
        if cumulative_prod >= final_gap:
            break
        daily_prod = int(single_shift_daily * (rate / 100.0))
        if cumulative_prod + daily_prod > final_gap:
            daily_prod = final_gap - cumulative_prod
        prod_list.append(daily_prod)
        cumulative_prod += daily_prod
        need_days += 1

    # 如果整个爬坡周期都不够，补满产天数
    if cumulative_prod < final_gap:
        remaining_gap = final_gap - cumulative_prod
        full_prod_days = math.ceil(remaining_gap / single_shift_daily)
        for i in range(full_prod_days):
            daily_prod = single_shift_daily
            if cumulative_prod + daily_prod > final_gap:
                daily_prod = final_gap - cumulative_prod
            prod_list.append(daily_prod)
            cumulative_prod += daily_prod
            need_days += 1

    return need_days, prod_list

def analyze_overtime_optimization(
    final_shift_total, existing_shift_count, single_shift_daily,
    production_end_date, rest_dates_set, full_date_list,
    target_shift_label="最后一个新增班组",
):
    """
    加班优化分析：针对最后一个新增班组的产量，判断是否可通过减少休息日完成
    """
    if final_shift_total <= 0:
        return False, "", []

    # 1. 筛选生产窗口期内的休息日（<=生产截止日）
    production_rest_dates = [d for d in rest_dates_set if d <= production_end_date and d in full_date_list]
    if len(production_rest_dates) == 0:
        return False, "", []

    # 2. 计算单个休息日改成工作日，现有班组能增加的总产能
    single_rest_day_capacity = existing_shift_count * single_shift_daily
    if single_rest_day_capacity <= 0:
        return False, "", []
    # 3. 计算需要最少多少个休息日
    need_days = math.ceil(final_shift_total / single_rest_day_capacity)

    if need_days > len(production_rest_dates):
        return False, "", []

    # 4. 优先推荐周六（先改周六，保留周日休息），按日期从晚到早排序
    saturday_rest_dates = [d for d in production_rest_dates if d.weekday() == 5]
    other_rest_dates = [d for d in production_rest_dates if d.weekday() != 5]
    sorted_rest_dates = sorted(saturday_rest_dates, reverse=True) + sorted(other_rest_dates, reverse=True)

    # 5. 取需要的日期
    selected_dates = sorted_rest_dates[:need_days]
    total_add_capacity = need_days * single_rest_day_capacity

    # 6. 生成建议文本（已删除：推荐修改的休息日）
    suggest_text = f"""
⚠️ 加班优化建议：当前{target_shift_label}的总产量，可通过前序班组减少休息日（加班）完成，可考虑减少该班组！
🔹 {target_shift_label}总产量：{final_shift_total:,}
🔹 单个休息日改为工作日，现有班组可增加产能：{single_rest_day_capacity:,}（现有{existing_shift_count}个班组满产）
🔹 最少需要减少休息日数量：{need_days}天
🔹 修改后可增加总产能：{total_add_capacity:,}，完全覆盖该班组产量
🔹 重要说明：减少休息日重新排产，依然完全遵从现有正排/倒排规则、爬坡规则，不会打乱原有逻辑
"""
    return True, suggest_text, selected_dates

def find_overtime_reduction_candidate(schedule_df):
    """优先检查新班组；没有新班组时检查最后一个已启用老班组。"""
    if schedule_df.empty:
        return 0, 0, ""
    date_cols = list(schedule_df.columns[2:])
    shift_rows = schedule_df[schedule_df["班组/指标"].astype(str).str.startswith("班组", na=False)]
    if shift_rows.empty:
        return 0, 0, ""

    new_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("新班组", na=False)]
    if not new_rows.empty:
        candidate_row = new_rows.iloc[-1]
        candidate_qty = sum(to_int_or_zero(candidate_row[col]) for col in date_cols)
        old_count = len(shift_rows[~shift_rows["班组/指标"].astype(str).str.contains("新班组", na=False)])
        return candidate_qty, max(old_count, 1), str(candidate_row["班组/指标"])

    old_rows = shift_rows[shift_rows["班组/指标"].astype(str).str.contains("老班组", na=False)]
    if len(old_rows) <= 1:
        return 0, 0, ""
    candidate_row = old_rows.iloc[-1]
    candidate_qty = sum(to_int_or_zero(candidate_row[col]) for col in date_cols)
    return candidate_qty, len(old_rows) - 1, str(candidate_row["班组/指标"])

# ------------------------------
# 三、核心排产引擎
# ------------------------------
def schedule_engine(
    process_name,
    total_demand, initial_stock, special_occupy,
    uph_base, work_hours, existing_shift_count,
    schedule_start_date, demand_end_date, lead_time_days,
    rest_dates_set,
    material_initial_stock, material_plan_df,
    selected_model, new_human_ratio,
    material_enabled=True,
    work_hours_plan_df=None,
    idle_convert_threshold_days=5,
):
    # 生产总目标
    production_target = max(total_demand + special_occupy - initial_stock, 0)
    original_production_target = int(production_target)
    material_limited_by_total = False
    material_shortage_qty = 0
    default_single_shift_daily = uph_base * work_hours

    if production_target == 0:
        return pd.DataFrame(), "生产目标为0，无需排产", "", production_target, 0, default_single_shift_daily, 0, demand_end_date, 0, []

    # 成品转化 Lead Time 按工作日计算，休息日不计入转换周期。
    production_end_date = get_previous_workday(demand_end_date, int(lead_time_days), rest_dates_set)
    full_date_list = generate_full_date_list(schedule_start_date, demand_end_date)
    total_days = len(full_date_list)
    if total_days == 0:
        return pd.DataFrame(), "错误：目标日期范围内无有效日期", "", production_target, 0, default_single_shift_daily, 0, demand_end_date, 0, []

    date_workday_flag = [d not in rest_dates_set for d in full_date_list]
    date_in_production_flag = [d <= production_end_date for d in full_date_list]
    date_to_idx = {d: i for i, d in enumerate(full_date_list)}
    total_workdays = sum(date_workday_flag)

    production_workday_indices = [i for i in range(total_days) if date_workday_flag[i] and date_in_production_flag[i]]
    reverse_production_workday_indices = production_workday_indices[::-1]
    production_workdays = len(production_workday_indices)

    if production_workdays == 0:
        return pd.DataFrame(), "错误：生产周期内无可用工作日", "", production_target, 0, default_single_shift_daily, 0, production_end_date, 0, []

    work_hours_plan, invalid_work_hours_rows = parse_work_hours_plan(work_hours_plan_df, int(work_hours))
    ignored_work_hours_dates = sorted(d for d in work_hours_plan if d not in set(full_date_list))
    daily_work_hours = [
        int(work_hours_plan.get(d, work_hours)) if date_workday_flag[idx] and date_in_production_flag[idx] else 0
        for idx, d in enumerate(full_date_list)
    ]
    daily_shift_capacity = [int(uph_base * hours) for hours in daily_work_hours]

    if material_enabled:
        material_plan, invalid_material_rows = parse_material_plan(material_plan_df)
        material_arrivals = [int(material_plan.get(d, 0)) for d in full_date_list]
        effective_material_arrivals = [0] * total_days
        for arrival_idx, qty in enumerate(material_arrivals):
            if int(qty) <= 0:
                continue
            available_date = get_next_workday(full_date_list[arrival_idx], 1, rest_dates_set)
            available_idx = date_to_idx.get(available_date)
            if available_idx is not None:
                effective_material_arrivals[available_idx] += int(qty)
        ignored_material_dates = sorted(d for d in material_plan if d not in set(full_date_list))
        material_total_available = max(0, int(material_initial_stock)) + sum(max(0, int(qty)) for qty in material_arrivals)
        if material_total_available < int(production_target):
            material_limited_by_total = True
            material_shortage_qty = max(0, original_production_target - int(material_total_available))
            production_target = int(material_total_available)
    else:
        invalid_material_rows = 0
        material_arrivals = [0] * total_days
        effective_material_arrivals = [0] * total_days
        ignored_material_dates = []
    daily_scheduled = [0] * total_days

    def material_available_at_day_start(day_idx, scheduled_snapshot):
        if not material_enabled:
            return 10**18
        # 当日可用物料 = 前一天累计物料gap + 前一天预计到料数量。
        available_qty = int(material_initial_stock) + int(effective_material_arrivals[day_idx])
        for idx in range(day_idx):
            available_qty = available_qty - int(scheduled_snapshot[idx]) + int(effective_material_arrivals[idx])
        return available_qty

    def available_material_for_day(day_idx):
        return max(0, material_available_at_day_start(day_idx, daily_scheduled) - int(daily_scheduled[day_idx]))

    def can_fully_produce_on_day(day_idx, desired_qty):
        return available_material_for_day(day_idx) >= int(desired_qty)

    def assign_production(shift, day_idx, desired_qty, remaining_qty, allow_partial=True):
        if allow_partial:
            actual_qty = min(int(desired_qty), int(remaining_qty), available_material_for_day(day_idx))
        elif int(remaining_qty) >= int(desired_qty) and available_material_for_day(day_idx) >= int(desired_qty):
            actual_qty = int(desired_qty)
        else:
            actual_qty = 0
        if actual_qty > 0:
            shift["daily_prod"][day_idx] = actual_qty
            daily_scheduled[day_idx] += actual_qty
        return actual_qty

    def init_shift(name, is_new, fill_idle=True):
        daily = []
        for idx in range(total_days):
            if fill_idle and date_workday_flag[idx] and date_in_production_flag[idx]:
                daily.append("当日放空")
            else:
                daily.append("")
        return {"name": name, "daily_prod": daily, "is_new": is_new}

    def count_shift_idle_days(shift, active_window_only=False):
        idle_days = 0
        if is_shift_empty(shift["daily_prod"]):
            return idle_days
        indices = production_workday_indices
        if active_window_only:
            active_indices = [
                idx for idx in production_workday_indices
                if isinstance(shift["daily_prod"][idx], (int, float)) and int(shift["daily_prod"][idx]) > 0
            ]
            if not active_indices:
                return idle_days
            indices = [idx for idx in production_workday_indices if active_indices[0] <= idx <= active_indices[-1]]
        # 只遍历生产窗口内的工作日；休息日不会进入 production_workday_indices。
        for idx in indices:
            day_capacity = int(daily_shift_capacity[idx])
            # 自定义为 0 工时的日期视同不可排产，不计入放空。
            if day_capacity <= 0:
                continue
            val = shift["daily_prod"][idx]
            # 空白、当日放空、未满产，均计为该班组工作日放空。
            if str(val).strip() == "":
                idle_days += 1
            elif val == "当日放空" or (isinstance(val, (int, float)) and int(val) < day_capacity):
                idle_days += 1
        return idle_days

    def count_old_shift_idle_days():
        idle_days = 0
        for shift in shifts_production:
            if shift["is_new"]:
                continue
            idle_days += count_shift_idle_days(shift, active_window_only=False)
        return idle_days

    def target_completion_day(scheduled_snapshot):
        produced = 0
        for day_idx, qty in enumerate(scheduled_snapshot):
            produced += int(qty)
            if produced >= int(production_target):
                return day_idx
        return production_workday_indices[-1]

    def max_consecutive_idle_before_completion(shift, scheduled_snapshot):
        completion_idx = target_completion_day(scheduled_snapshot)
        max_idle = 0
        current_idle = 0
        for day_idx in production_workday_indices:
            if day_idx > completion_idx:
                break
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            val = shift["daily_prod"][day_idx]
            is_idle = str(val).strip() == "" or val == "当日放空"
            is_partial = isinstance(val, (int, float)) and int(val) < day_capacity
            if is_idle or is_partial:
                current_idle += 1
                max_idle = max(max_idle, current_idle)
            else:
                current_idle = 0
        return max_idle

    def count_idle_before_completion(shift, scheduled_snapshot):
        completion_idx = target_completion_day(scheduled_snapshot)
        idle_days = 0
        for day_idx in production_workday_indices:
            if day_idx > completion_idx:
                break
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            val = shift["daily_prod"][day_idx]
            is_idle = str(val).strip() == "" or val == "当日放空"
            is_partial = isinstance(val, (int, float)) and int(val) < day_capacity
            if is_idle or is_partial:
                idle_days += 1
        return idle_days

    def count_capacity_workdays_between(start_idx, end_idx):
        if start_idx is None or end_idx is None or end_idx < start_idx:
            return 0
        return sum(
            1
            for day_idx in production_workday_indices
            if start_idx <= day_idx <= end_idx and int(daily_shift_capacity[day_idx]) > 0
        )

    def ideal_completion_day_for_shift_qty(target_qty, start_idx):
        produced = 0
        for day_idx in production_workday_indices:
            if day_idx < start_idx:
                continue
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            produced += min(day_capacity, int(target_qty) - produced)
            if produced >= int(target_qty):
                return day_idx
        return None

    def count_last_old_shift_delay_days(shift):
        target_qty = sum(numeric_prod(shift["daily_prod"][day_idx]) for day_idx in range(total_days))
        if target_qty <= 0 or not production_workday_indices:
            return 0
        produced_indices = [
            day_idx
            for day_idx in production_workday_indices
            if numeric_prod(shift["daily_prod"][day_idx]) > 0
        ]
        if not produced_indices:
            return 0
        # 最后一个老班组只判断它自己承担的补量段：
        # 从该班组首次实际生产日开始，比较“物料满足时应完成的工作日跨度”和实际完成跨度。
        ideal_start_idx = produced_indices[0]
        ideal_completion_idx = ideal_completion_day_for_shift_qty(target_qty, ideal_start_idx)
        actual_completion_idx = produced_indices[-1]
        if ideal_completion_idx is None:
            return count_capacity_workdays_between(ideal_start_idx, actual_completion_idx)
        ideal_days = count_capacity_workdays_between(ideal_start_idx, ideal_completion_idx)
        actual_days = count_capacity_workdays_between(ideal_start_idx, actual_completion_idx)
        return max(0, actual_days - ideal_days)

    def can_place_sequence(start_pos, prod_sequence, scheduled_snapshot):
        simulated = scheduled_snapshot[:]
        for offset, prod in enumerate(prod_sequence):
            seq_pos = start_pos + offset
            if seq_pos >= len(production_workday_indices):
                return False
            day_idx = production_workday_indices[seq_pos]
            available_qty = max(0, material_available_at_day_start(day_idx, simulated) - int(simulated[day_idx]))
            if available_qty < int(prod):
                return False
            simulated[day_idx] += int(prod)
        return True

    def place_sequence(shift, start_pos, prod_sequence):
        produced = 0
        for offset, prod in enumerate(prod_sequence):
            day_idx = production_workday_indices[start_pos + offset]
            shift["daily_prod"][day_idx] = int(prod)
            daily_scheduled[day_idx] += int(prod)
            produced += int(prod)
        return produced

    def build_one_shift_sequence(target_qty, start_pos=0, max_days=None):
        ramp_curve = RAMP_DATA[selected_model][new_human_ratio]
        seq = []
        produced = 0
        max_days = production_workdays if max_days is None else int(max_days)
        for workday_seq in range(max_days):
            seq_pos = start_pos + workday_seq
            if seq_pos >= len(production_workday_indices):
                break
            day_idx = production_workday_indices[seq_pos]
            rate = ramp_curve[workday_seq] if workday_seq < len(ramp_curve) else 100
            qty = int(daily_shift_capacity[day_idx] * (rate / 100.0))
            qty = min(qty, int(target_qty) - produced)
            if qty <= 0:
                break
            seq.append(qty)
            produced += qty
            if produced >= int(target_qty):
                break
        return seq

    def calc_material_gap_row(scheduled_snapshot):
        gap_row = []
        for idx, qty in enumerate(scheduled_snapshot):
            # 当天排产只能消耗当日可用物料；当天到料按 T+1 进入下一天可用量。
            gap_row.append(max(0, material_available_at_day_start(idx, scheduled_snapshot) - int(qty)))
        return gap_row

    def calc_material_available_row(scheduled_snapshot):
        return [material_available_at_day_start(idx, scheduled_snapshot) for idx in range(total_days)]

    def add_new_reverse_shifts(target_qty, reset_existing=True):
        nonlocal daily_scheduled, shifts_production
        if reset_existing:
            daily_scheduled = [0] * total_days
            shifts_production = []
        remaining = int(target_qty)
        new_shift_total = 0
        shift_no = len(shifts_production) + 1

        while remaining > 0 and shift_no <= 50:
            best_start = None
            best_sequence = []
            best_total = 0
            for start_pos in range(len(production_workday_indices)):
                max_days = len(production_workday_indices) - start_pos
                prod_sequence = build_one_shift_sequence(remaining, start_pos=start_pos, max_days=max_days)
                if not prod_sequence:
                    continue
                produced = sum(prod_sequence)
                if produced <= 0:
                    continue
                if not can_place_sequence(start_pos, prod_sequence, daily_scheduled):
                    continue
                if produced > best_total or (produced == best_total and (best_start is None or start_pos > best_start)):
                    best_start = start_pos
                    best_sequence = prod_sequence
                    best_total = produced

            if best_start is None:
                break

            shift = init_shift(f"班组{shift_no}(新班组-连续正排)", True, fill_idle=False)
            produced = place_sequence(shift, best_start, best_sequence)
            shifts_production.append(shift)
            remaining -= produced
            new_shift_total += produced
            shift_no += 1

        return remaining, new_shift_total

    def run_old_forward_schedule(old_shift_count, target_qty):
        nonlocal daily_scheduled
        daily_scheduled = [0] * total_days
        candidate_shifts = [init_shift(f"班组{i+1}(老班组-连续正排)", False, fill_idle=False) for i in range(old_shift_count)]
        remaining = int(target_qty)

        def schedule_one_slot(shift_idx, day_idx):
            nonlocal remaining
            if remaining <= 0:
                return 0
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                return 0
            available_qty = available_material_for_day(day_idx)
            if available_qty <= 0:
                candidate_shifts[shift_idx]["daily_prod"][day_idx] = "当日放空"
                return 0
            if material_enabled:
                prod = min(day_capacity, remaining)
                actual = assign_production(
                    candidate_shifts[shift_idx],
                    day_idx,
                    prod,
                    remaining,
                    allow_partial=True,
                )
            else:
                actual = min(day_capacity, remaining)
                candidate_shifts[shift_idx]["daily_prod"][day_idx] = int(actual)
                daily_scheduled[day_idx] += int(actual)
            remaining -= actual
            if actual <= 0 and remaining > 0:
                candidate_shifts[shift_idx]["daily_prod"][day_idx] = "当日放空"
            return actual

        # 老班组按班组优先级连续正排：先排满班组1，再排班组2，最后依次处理后续老班组。
        # 启用物料时，每个班组仍按时间顺序消耗当日可用物料，物料不足才形成放空或尾数。
        for shift_idx in range(old_shift_count):
            if remaining <= 0:
                break
            for day_idx in production_workday_indices:
                if remaining <= 0:
                    break
                schedule_one_slot(shift_idx, day_idx)

        return candidate_shifts, daily_scheduled[:], remaining

    def numeric_prod(value):
        return int(value) if isinstance(value, (int, float)) else 0

    def set_shift_day_value(shift, day_idx, qty):
        if qty > 0:
            shift["daily_prod"][day_idx] = int(qty)
        elif date_workday_flag[day_idx] and date_in_production_flag[day_idx] and int(daily_shift_capacity[day_idx]) > 0:
            shift["daily_prod"][day_idx] = "当日放空"
        else:
            shift["daily_prod"][day_idx] = ""

    def prioritize_old_shifts_and_cap_material():
        nonlocal daily_scheduled, remaining_demand

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        new_shifts = [shift for shift in shifts_production if shift["is_new"]]

        # 物料硬约束：任何一天总排产不能超过当天可用物料，超出部分优先从新班组扣回。
        # 不再把新班组爬坡产量搬回老班组，否则会破坏新班组启动日必须产出的规则。
        available_qty = int(material_initial_stock) if material_enabled else 10**18
        for day_idx in range(total_days):
            day_total = sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            allowed_qty = max(0, int(available_qty))
            excess = max(0, day_total - allowed_qty)
            if excess > 0:
                for shift in list(reversed(new_shifts)) + list(reversed(old_shifts)):
                    if excess <= 0:
                        break
                    current_qty = numeric_prod(shift["daily_prod"][day_idx])
                    if current_qty <= 0:
                        continue
                    reduce_qty = min(current_qty, excess)
                    current_qty -= reduce_qty
                    excess -= reduce_qty
                    remaining_demand += reduce_qty
                    set_shift_day_value(shift, day_idx, current_qty)
                day_total = sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            if material_enabled:
                available_qty = max(0, available_qty - day_total) + int(material_arrivals[day_idx])

        daily_scheduled = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            for day_idx in range(total_days)
        ]

    def compact_old_shift_usage_by_deferring():
        nonlocal daily_scheduled

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        if len(old_shifts) <= 1:
            return

        workday_pos_by_idx = {day_idx: pos for pos, day_idx in enumerate(production_workday_indices)}

        for src_shift_idx in range(len(old_shifts) - 1, 0, -1):
            src_shift = old_shifts[src_shift_idx]
            for src_day_idx in production_workday_indices:
                src_qty = numeric_prod(src_shift["daily_prod"][src_day_idx])
                if src_qty <= 0:
                    continue
                src_capacity = int(daily_shift_capacity[src_day_idx])
                if src_capacity <= 0:
                    continue
                if src_qty >= src_capacity:
                    continue

                remaining_to_move = src_qty
                placements = []
                src_pos = workday_pos_by_idx[src_day_idx]

                for target_day_idx in production_workday_indices[src_pos + 1:]:
                    if remaining_to_move <= 0:
                        break
                    target_capacity = int(daily_shift_capacity[target_day_idx])
                    if target_capacity <= 0:
                        continue
                    for target_shift_idx in range(src_shift_idx):
                        if remaining_to_move <= 0:
                            break
                        target_shift = old_shifts[target_shift_idx]
                        target_qty = numeric_prod(target_shift["daily_prod"][target_day_idx])
                        target_room = max(0, target_capacity - target_qty)
                        if target_room <= 0:
                            continue
                        move_qty = min(remaining_to_move, target_room)
                        placements.append((target_shift, target_day_idx, move_qty))
                        remaining_to_move -= move_qty

                if remaining_to_move > 0:
                    continue

                src_shift["daily_prod"][src_day_idx] = ""
                for target_shift, target_day_idx, move_qty in placements:
                    target_qty = numeric_prod(target_shift["daily_prod"][target_day_idx])
                    set_shift_day_value(target_shift, target_day_idx, target_qty + move_qty)

        daily_scheduled = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            for day_idx in range(total_days)
        ]

    def top_up_unfinished_target_with_old_shifts():
        nonlocal daily_scheduled, remaining_demand

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        produced_qty = sum(int(v) for v in daily_scheduled)
        remaining_qty = max(0, int(production_target) - int(produced_qty))
        if remaining_qty <= 0 or not old_shifts:
            remaining_demand = remaining_qty
            return 0

        filled_qty = 0
        for day_idx in reversed(production_workday_indices):
            if remaining_qty <= 0:
                break
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            available_qty = available_material_for_day(day_idx)
            if available_qty <= 0:
                continue
            for shift in old_shifts:
                if remaining_qty <= 0 or available_qty <= 0:
                    break
                current_qty = numeric_prod(shift["daily_prod"][day_idx])
                room = max(0, day_capacity - current_qty)
                if room <= 0:
                    continue
                add_qty = min(room, available_qty, remaining_qty)
                set_shift_day_value(shift, day_idx, current_qty + add_qty)
                daily_scheduled[day_idx] += int(add_qty)
                available_qty -= int(add_qty)
                remaining_qty -= int(add_qty)
                filled_qty += int(add_qty)

        remaining_demand = remaining_qty
        return filled_qty

    def fill_old_shifts_forward_to_target():
        nonlocal daily_scheduled, remaining_demand

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        remaining_qty = max(0, int(production_target) - sum(int(v) for v in daily_scheduled))
        if remaining_qty <= 0 or not old_shifts:
            remaining_demand = remaining_qty
            return 0

        filled_qty = 0
        for day_idx in production_workday_indices:
            if remaining_qty <= 0:
                break
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            available_qty = available_material_for_day(day_idx)
            if available_qty <= 0:
                continue
            for shift in old_shifts:
                if remaining_qty <= 0 or available_qty <= 0:
                    break
                current_qty = numeric_prod(shift["daily_prod"][day_idx])
                room = max(0, day_capacity - current_qty)
                if room <= 0:
                    continue
                add_qty = min(room, available_qty, remaining_qty)
                set_shift_day_value(shift, day_idx, current_qty + add_qty)
                daily_scheduled[day_idx] += int(add_qty)
                available_qty -= int(add_qty)
                remaining_qty -= int(add_qty)
                filled_qty += int(add_qty)

        remaining_demand = remaining_qty
        return filled_qty

    def pull_later_production_into_old_idle_slots():
        nonlocal daily_scheduled

        if not material_enabled:
            return 0

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        if not old_shifts:
            return 0

        daily_scheduled = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            for day_idx in range(total_days)
        ]
        moved_total = 0

        def material_plan_is_feasible(candidate_daily):
            for check_idx in range(total_days):
                if int(candidate_daily[check_idx]) > material_available_at_day_start(check_idx, candidate_daily):
                    return False
            return True

        def max_feasible_move(target_day_idx, source_day_idx, limit_qty):
            low = 0
            high = int(limit_qty)
            while low < high:
                mid = (low + high + 1) // 2
                candidate_daily = daily_scheduled[:]
                candidate_daily[target_day_idx] += mid
                candidate_daily[source_day_idx] -= mid
                if material_plan_is_feasible(candidate_daily):
                    low = mid
                else:
                    high = mid - 1
            return low

        def source_candidates(target_shift, target_day_idx):
            candidates = []
            target_shift_pos = old_shifts.index(target_shift)

            for other_shift in old_shifts[target_shift_pos + 1:]:
                if numeric_prod(other_shift["daily_prod"][target_day_idx]) > 0:
                    candidates.append((other_shift, target_day_idx))

            for source_day_idx in production_workday_indices:
                if source_day_idx <= target_day_idx:
                    continue
                if numeric_prod(target_shift["daily_prod"][source_day_idx]) > 0:
                    candidates.append((target_shift, source_day_idx))

            for other_shift in old_shifts[target_shift_pos + 1:]:
                for source_day_idx in production_workday_indices:
                    if source_day_idx <= target_day_idx:
                        continue
                    if numeric_prod(other_shift["daily_prod"][source_day_idx]) > 0:
                        candidates.append((other_shift, source_day_idx))

            return candidates

        moved = True
        while moved:
            moved = False
            for target_day_idx in production_workday_indices:
                day_capacity = int(daily_shift_capacity[target_day_idx])
                if day_capacity <= 0:
                    continue

                for target_shift in old_shifts:
                    target_qty = numeric_prod(target_shift["daily_prod"][target_day_idx])
                    target_room = max(0, day_capacity - target_qty)
                    if target_room <= 0:
                        continue

                    for source_shift, source_day_idx in source_candidates(target_shift, target_day_idx):
                        if target_room <= 0:
                            break
                        source_qty = numeric_prod(source_shift["daily_prod"][source_day_idx])
                        if source_qty <= 0:
                            continue
                        move_qty = max_feasible_move(target_day_idx, source_day_idx, min(target_room, source_qty))
                        if move_qty <= 0:
                            continue

                        target_qty += move_qty
                        source_qty -= move_qty
                        set_shift_day_value(target_shift, target_day_idx, target_qty)
                        if source_shift["is_new"] and source_qty <= 0:
                            source_shift["daily_prod"][source_day_idx] = ""
                        else:
                            set_shift_day_value(source_shift, source_day_idx, source_qty)
                        daily_scheduled[target_day_idx] += move_qty
                        daily_scheduled[source_day_idx] -= move_qty
                        target_room -= move_qty
                        moved_total += move_qty
                        moved = True

        return moved_total

    def rebuild_old_shifts_by_priority():
        nonlocal daily_scheduled, remaining_demand

        if not material_enabled:
            return 0

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        if not old_shifts:
            return 0

        new_daily = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production if shift["is_new"])
            for day_idx in range(total_days)
        ]
        new_total = sum(int(v) for v in new_daily)
        old_target = max(0, int(production_target) - int(new_total))

        for shift in old_shifts:
            shift["daily_prod"] = [""] * total_days

        daily_scheduled = new_daily[:]
        remaining_old_qty = int(old_target)
        scheduled_old_qty = 0

        for shift in old_shifts:
            if remaining_old_qty <= 0:
                break
            for day_idx in production_workday_indices:
                if remaining_old_qty <= 0:
                    break
                day_capacity = int(daily_shift_capacity[day_idx])
                if day_capacity <= 0:
                    continue
                available_qty = available_material_for_day(day_idx)
                if available_qty <= 0:
                    continue
                prod_qty = min(day_capacity, available_qty, remaining_old_qty)
                if prod_qty <= 0:
                    continue
                set_shift_day_value(shift, day_idx, prod_qty)
                daily_scheduled[day_idx] += int(prod_qty)
                remaining_old_qty -= int(prod_qty)
                scheduled_old_qty += int(prod_qty)

        remaining_demand = max(0, int(production_target) - int(new_total) - int(scheduled_old_qty))
        return scheduled_old_qty

    def normalize_shift_idle_display():
        daily_scheduled_snapshot = [
            sum(numeric_prod(item["daily_prod"][day_idx]) for item in shifts_production)
            for day_idx in range(total_days)
        ]
        for shift in shifts_production:
            if is_shift_empty(shift["daily_prod"]):
                continue
            if shift["is_new"]:
                continue
            produced_indices = [
                idx for idx, val in enumerate(shift["daily_prod"])
                if isinstance(val, (int, float)) and int(val) > 0
            ]
            last_prod_idx = produced_indices[-1] if produced_indices else -1
            for day_idx in production_workday_indices:
                if int(daily_shift_capacity[day_idx]) <= 0:
                    continue
                if last_prod_idx >= 0 and day_idx > last_prod_idx:
                    continue
                if str(shift["daily_prod"][day_idx]).strip() == "":
                    available_qty = max(
                        0,
                        material_available_at_day_start(day_idx, daily_scheduled_snapshot)
                        - int(daily_scheduled_snapshot[day_idx])
                    )
                    if available_qty <= 0:
                        shift["daily_prod"][day_idx] = "当日放空"

    def convert_non_continuous_old_shifts_to_new():
        nonlocal daily_scheduled, remaining_demand, final_shift_total, run_mode, message

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        first_convert_idx = None
        if material_enabled:
            current_material_gap = calc_material_gap_row(daily_scheduled)
            if current_material_gap and min(current_material_gap) > 0:
                return 0, 0, 0
        for idx, shift in enumerate(old_shifts):
            if material_enabled and idx == len(old_shifts) - 1:
                continue
            idle_days = count_idle_before_completion(shift, daily_scheduled)
            if idle_days > idle_convert_threshold_days:
                first_convert_idx = idx
                break
        if first_convert_idx is None:
            return 0, 0, 0

        converted_shifts = old_shifts[first_convert_idx:]
        converted_qty = sum(
            numeric_prod(shift["daily_prod"][day_idx])
            for shift in converted_shifts
            for day_idx in range(total_days)
        )
        if converted_qty <= 0:
            return 0, 0, 0

        converted_ids = {id(shift) for shift in converted_shifts}
        shifts_production[:] = [shift for shift in shifts_production if id(shift) not in converted_ids]
        daily_scheduled = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            for day_idx in range(total_days)
        ]

        remaining_after_new, new_total = add_new_reverse_shifts(converted_qty, reset_existing=False)
        final_shift_total += new_total
        remaining_demand += remaining_after_new

        converted_count = len(converted_shifts)
        if new_total > 0:
            run_mode = "场景二/四：物料连续缺口，老班组顺序正排 + 新班组爬坡正排模式"
            message = (
                f"✅ 排产完成 | {run_mode} | 第{first_convert_idx + 1}个及后续老班组"
                f"生产完成前累计放空超过{idle_convert_threshold_days}天，已按新班组爬坡正排处理"
            )
        if remaining_after_new > 0:
            message = (
                f"⚠️ 排产未完全覆盖 | {run_mode} | 第{first_convert_idx + 1}个及后续老班组"
                f"生产完成前累计放空超过{idle_convert_threshold_days}天，转新班组后仍有{remaining_after_new:,}件未排完"
            )
        return converted_count, converted_qty, remaining_after_new

    def rebuild_new_shifts_to_target():
        nonlocal remaining_demand, final_shift_total, run_mode, message

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        old_daily = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in old_shifts)
            for day_idx in range(total_days)
        ]
        shifts_production[:] = old_shifts
        daily_scheduled[:] = old_daily
        final_shift_total = 0
        fill_old_shifts_forward_to_target()
        old_total = sum(int(v) for v in daily_scheduled)
        new_target_qty = max(0, int(production_target) - int(old_total))

        if new_target_qty <= 0:
            remaining_demand = 0
            return 0, 0

        remaining_after_new, new_total = add_new_reverse_shifts(new_target_qty, reset_existing=False)
        final_shift_total = new_total
        remaining_demand = remaining_after_new

        if new_total > 0:
            run_mode = "场景二/四：物料连续缺口，老班组顺序正排 + 新班组爬坡正排模式"
            if remaining_after_new > 0:
                message = (
                    f"⚠️ 排产未完全覆盖 | {run_mode} | 已按新班组补排{new_total:,}件，"
                    f"仍有{remaining_after_new:,}件受物料交期或生产窗口限制未排完"
                )
            else:
                message = f"✅ 排产完成 | {run_mode} | 已用新班组补足剩余需求"

        return remaining_after_new, new_total

    # 产能计算
    def rebalance_new_output_to_last_old_shift_if_within_delay():
        nonlocal daily_scheduled, remaining_demand, final_shift_total, run_mode, message

        if not material_enabled:
            return False

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        new_shifts = [shift for shift in shifts_production if shift["is_new"]]
        if not old_shifts or not new_shifts:
            return False

        last_old_shift = old_shifts[-1]
        target_qty = sum(numeric_prod(last_old_shift["daily_prod"][day_idx]) for day_idx in range(total_days))
        target_qty += sum(
            numeric_prod(shift["daily_prod"][day_idx])
            for shift in new_shifts
            for day_idx in range(total_days)
        )
        if target_qty <= 0:
            return False

        simulated = [
            sum(
                numeric_prod(shift["daily_prod"][day_idx])
                for shift in old_shifts
                if id(shift) != id(last_old_shift)
            )
            for day_idx in range(total_days)
        ]
        candidate_daily = [""] * total_days
        remaining = int(target_qty)

        for day_idx in production_workday_indices:
            if remaining <= 0:
                break
            day_capacity = int(daily_shift_capacity[day_idx])
            if day_capacity <= 0:
                continue
            available_qty = max(0, material_available_at_day_start(day_idx, simulated) - int(simulated[day_idx]))
            actual = min(day_capacity, available_qty, remaining)
            if actual <= 0:
                continue
            candidate_daily[day_idx] = int(actual)
            simulated[day_idx] += int(actual)
            remaining -= int(actual)

        if remaining > 0:
            return False

        produced_indices = [
            day_idx for day_idx in production_workday_indices
            if numeric_prod(candidate_daily[day_idx]) > 0
        ]
        if not produced_indices:
            return False
        ideal_start_idx = produced_indices[0]
        ideal_completion_idx = ideal_completion_day_for_shift_qty(target_qty, ideal_start_idx)
        actual_completion_idx = produced_indices[-1]
        if ideal_completion_idx is None:
            return False
        ideal_days = count_capacity_workdays_between(ideal_start_idx, ideal_completion_idx)
        actual_days = count_capacity_workdays_between(ideal_start_idx, actual_completion_idx)
        delay_days = max(0, actual_days - ideal_days)
        if delay_days > idle_convert_threshold_days:
            return False

        last_old_shift["daily_prod"] = candidate_daily
        new_ids = {id(shift) for shift in new_shifts}
        shifts_production[:] = [shift for shift in shifts_production if id(shift) not in new_ids]
        daily_scheduled = simulated
        remaining_demand = 0
        final_shift_total = 0
        run_mode = "场景三：物料间歇缺口，保留老班组放空并后续补产"
        message = f"✅ 排产完成 | {run_mode} | 最后一个老班组缺料放空{delay_days}天，未超过{idle_convert_threshold_days}天，保留老班组"
        return True

    def convert_last_old_shift_to_new_if_delay_exceeds():
        nonlocal daily_scheduled, remaining_demand, final_shift_total, run_mode, message

        if not material_enabled:
            return 0, 0, 0

        old_shifts = [shift for shift in shifts_production if not shift["is_new"]]
        new_shifts = [shift for shift in shifts_production if shift["is_new"]]
        if not old_shifts or new_shifts:
            return 0, 0, 0

        last_old_shift = old_shifts[-1]
        target_qty = sum(numeric_prod(last_old_shift["daily_prod"][day_idx]) for day_idx in range(total_days))
        if target_qty <= 0:
            return 0, 0, 0

        delay_days = count_last_old_shift_delay_days(last_old_shift)
        if delay_days <= idle_convert_threshold_days:
            return 0, target_qty, delay_days

        shifts_production[:] = [shift for shift in shifts_production if id(shift) != id(last_old_shift)]
        daily_scheduled = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production)
            for day_idx in range(total_days)
        ]

        remaining_after_new, new_total = add_new_reverse_shifts(target_qty, reset_existing=False)
        final_shift_total += new_total
        remaining_demand += remaining_after_new
        run_mode = "场景二/四：物料连续缺口，老班组顺序正排 + 新班组爬坡正排模式"
        if remaining_after_new > 0:
            message = (
                f"⚠️ 排产未完全覆盖 | {run_mode} | 最后一个老班组因缺料放空超过{idle_convert_threshold_days}天，"
                f"转新班组后仍有{remaining_after_new:,}件未排完"
            )
        else:
            message = (
                f"✅ 排产完成 | {run_mode} | 最后一个老班组因缺料放空超过{idle_convert_threshold_days}天，"
                f"已按新班组爬坡正排处理"
            )
        return 1, new_total, delay_days

    capacity_1_shift_full = sum(int(daily_shift_capacity[idx]) for idx in production_workday_indices)
    total_exist_capacity = capacity_1_shift_full * existing_shift_count
    demand_gap = production_target - total_exist_capacity
    small_gap_threshold = 2 * default_single_shift_daily

    # ============================
    # 初始化班组：后续按实际启用班组数生成，只保留覆盖目标所需班组
    # ============================
    shifts_production = []
    run_mode = ""

    # ============================
    # 场景1：现有老班组产能可覆盖目标
    # ============================
    final_shift_total = 0
    old_material_gap_row = None
    if demand_gap <= 0:
        run_mode = "老班组优选模式（目标可覆盖，正排逻辑）"
        selected_old_count = existing_shift_count
        selected_remaining = production_target
        selected_shifts = []
        selected_daily = [0] * total_days

        for candidate_count in range(1, existing_shift_count + 1):
            if candidate_count * capacity_1_shift_full < production_target:
                continue
            candidate_shifts, candidate_daily, candidate_remaining = run_old_forward_schedule(candidate_count, production_target)
            if candidate_remaining <= 0:
                selected_old_count = candidate_count
                selected_remaining = 0
                selected_shifts = candidate_shifts
                selected_daily = candidate_daily
                break
            if not selected_shifts or candidate_remaining < selected_remaining:
                selected_old_count = candidate_count
                selected_remaining = candidate_remaining
                selected_shifts = candidate_shifts
                selected_daily = candidate_daily

        if not selected_shifts:
            selected_shifts, selected_daily, selected_remaining = run_old_forward_schedule(existing_shift_count, production_target)

        shifts_production = selected_shifts
        daily_scheduled = selected_daily
        remaining_demand = selected_remaining

        old_material_gap_row = calc_material_gap_row(daily_scheduled)
        idle_days = count_old_shift_idle_days()

        if remaining_demand > 0:
            message = f"⚠️ 排产未完全覆盖 | {run_mode} | 受物料交期或生产窗口约束，仍有{remaining_demand:,}件未排完"
        elif final_shift_total > 0:
            message = f"✅ 排产完成 | {run_mode} | 老班组正排存在物料缺口，已启用新班组连续正排补足"
        else:
            reduced_count = existing_shift_count - selected_old_count
            message = f"✅ 排产完成 | {run_mode} | 启用老班组{selected_old_count}个，减少{reduced_count}个 | 老班组工作日放空{idle_days}天"

    # ============================
    # 场景2：产能不足
    # ============================
    else:
        run_mode = "新增班组模式（产能不足，新增班组严格遵循爬坡规则连续正排）"
        shifts_production, daily_scheduled, remaining_demand = run_old_forward_schedule(existing_shift_count, production_target)

        final_shift_total = 0
        old_material_gap_row = calc_material_gap_row(daily_scheduled)
        idle_days = count_old_shift_idle_days()
        old_shift_count_after = len(shifts_production)
        remaining_demand, final_shift_total = add_new_reverse_shifts(remaining_demand, reset_existing=False) if remaining_demand > 0 else (0, 0)
        if old_shift_count_after and final_shift_total > 0:
            run_mode = "老班组正排 + 新班组连续正排模式"
        if remaining_demand > 0 and final_shift_total <= 0:
            message = f"⚠️ 排产未完全覆盖 | {run_mode} | 受物料交期约束，仍有{remaining_demand:,}件未排完"
        else:
            message = f"✅ 排产完成 | {run_mode} | 启用班组总数：{len(shifts_production)}个 | 生产必须完成截止日：{production_end_date.month}月{production_end_date.day}日"

    prioritize_old_shifts_and_cap_material()
    if material_enabled:
        if demand_gap <= 0:
            fill_old_shifts_forward_to_target()
            prioritize_old_shifts_and_cap_material()
            converted_count, _, _ = convert_non_continuous_old_shifts_to_new()
            prioritize_old_shifts_and_cap_material()
            if converted_count > 0:
                rebuild_new_shifts_to_target()
                prioritize_old_shifts_and_cap_material()
                rebalance_new_output_to_last_old_shift_if_within_delay()
                prioritize_old_shifts_and_cap_material()
            convert_last_old_shift_to_new_if_delay_exceeds()
            prioritize_old_shifts_and_cap_material()
        else:
            top_up_unfinished_target_with_old_shifts()
            prioritize_old_shifts_and_cap_material()
            convert_non_continuous_old_shifts_to_new()
            prioritize_old_shifts_and_cap_material()
            rebuild_new_shifts_to_target()
            prioritize_old_shifts_and_cap_material()
            rebalance_new_output_to_last_old_shift_if_within_delay()
            prioritize_old_shifts_and_cap_material()
            convert_last_old_shift_to_new_if_delay_exceeds()
            prioritize_old_shifts_and_cap_material()
        pull_later_production_into_old_idle_slots()
        prioritize_old_shifts_and_cap_material()
        fill_old_shifts_forward_to_target()
        prioritize_old_shifts_and_cap_material()
        rebuild_old_shifts_by_priority()
        prioritize_old_shifts_and_cap_material()
    normalize_shift_idle_display()

    # ============================
    # 统计行
    # ============================
    daily_total = [0] * total_days
    for day_idx in range(total_days):
        day_sum = 0
        for shift in shifts_production:
            val = shift["daily_prod"][day_idx]
            if isinstance(val, (int, float)):
                day_sum += int(val)
        daily_total[day_idx] = day_sum

    if material_enabled and any(not is_shift_empty(shift["daily_prod"]) for shift in shifts_production if shift["is_new"]):
        old_daily_total = [
            sum(numeric_prod(shift["daily_prod"][day_idx]) for shift in shifts_production if not shift["is_new"])
            for day_idx in range(total_days)
        ]
        old_material_gap_row = calc_material_gap_row(old_daily_total)

    material_gap_row = []
    material_usable_row = []
    if material_enabled:
        for day_idx in range(total_days):
            available_qty = max(0, material_available_at_day_start(day_idx, daily_total))
            material_usable_row.append(int(available_qty))
            material_gap_row.append(max(0, int(available_qty) - int(daily_total[day_idx])))

    # 累计产量
    cumulative_row = [int(initial_stock)]
    current_cum = int(initial_stock)
    for val in daily_total:
        current_cum += val
        cumulative_row.append(current_cum)

    # 成品转化累计：Lead Time 按工作日偏移，T 日生产在 T+Lead Time 个工作日后转化。
    convert_row = [int(initial_stock)]
    convert_additions = [0] * total_days
    for prod_idx, qty in enumerate(daily_total):
        complete_date = get_next_workday(full_date_list[prod_idx], int(lead_time_days), rest_dates_set)
        complete_idx = date_to_idx.get(complete_date)
        if complete_idx is not None:
            convert_additions[complete_idx] += int(qty)

    convert_cum = int(initial_stock)
    for day_idx in range(total_days):
        convert_cum += int(convert_additions[day_idx])
        convert_row.append(int(convert_cum))

    # 隐藏空班组
    final_shifts = []
    for shift in shifts_production:
        if not is_shift_empty(shift["daily_prod"]):
            final_shifts.append(shift)
    for display_no, shift in enumerate(final_shifts, start=1):
        shift_type = "新班组-连续正排" if shift["is_new"] else "老班组-连续正排"
        shift["name"] = f"班组{display_no}({shift_type})"

    # 新班组启动前不视为放空；第一天实际生产之前保持空白。
    for shift in final_shifts:
        if not shift["is_new"]:
            continue
        produced_indices = [
            idx for idx, val in enumerate(shift["daily_prod"])
            if isinstance(val, (int, float)) and int(val) > 0
        ]
        if not produced_indices:
            continue
        first_prod_idx = produced_indices[0]
        for day_idx in range(first_prod_idx):
            if shift["daily_prod"][day_idx] == "当日放空":
                shift["daily_prod"][day_idx] = ""

    # 显示层标记：班组后续不再生产时，只在第一个有效工作日标记释放。
    for shift in final_shifts:
        if shift["is_new"]:
            continue
        produced_indices = [
            idx for idx, val in enumerate(shift["daily_prod"])
            if isinstance(val, (int, float)) and int(val) > 0
        ]
        if not produced_indices:
            continue
        last_prod_idx = produced_indices[-1]
        release_marked = False
        for day_idx in range(last_prod_idx + 1, total_days):
            if not date_workday_flag[day_idx] or int(daily_shift_capacity[day_idx]) <= 0:
                continue
            current_val = shift["daily_prod"][day_idx]
            if not release_marked and (str(current_val).strip() == "" or current_val == "当日放空"):
                shift["daily_prod"][day_idx] = "班组释放"
                release_marked = True
            elif release_marked and current_val == "当日放空":
                shift["daily_prod"][day_idx] = ""

    final_shift_total = sum(
        numeric_prod(shift["daily_prod"][day_idx])
        for shift in final_shifts
        if shift["is_new"]
        for day_idx in range(total_days)
    )

    # 表格
    columns = ["班组/指标", "期初数据"]
    for d in full_date_list:
        week_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][d.weekday()]
        columns.append(f"{d.month}月{d.day}日\n{week_name}")

    rows = []
    rows.append(["单班组单日工时", ""] + daily_work_hours)
    if material_enabled:
        rows.append(["预计到料数量", int(material_initial_stock)] + material_arrivals)
        rows.append(["当日可用物料", int(material_initial_stock)] + material_usable_row)
    rows.append(["特殊占用（需生产）", int(special_occupy)] + [""] * total_days)
    old_shifts = [shift for shift in final_shifts if not shift["is_new"]]
    new_shifts = [shift for shift in final_shifts if shift["is_new"]]
    for shift in old_shifts:
        rows.append([shift["name"], ""] + shift["daily_prod"])
    if material_enabled and new_shifts and old_material_gap_row is not None:
        rows.append(["老班组后累计物料gap", ""] + old_material_gap_row)
    for shift in new_shifts:
        rows.append([shift["name"], ""] + shift["daily_prod"])
    if material_enabled:
        rows.append(["累计物料gap", int(material_initial_stock)] + material_gap_row)
    rows.append(["累计产量", int(initial_stock)] + cumulative_row[1:])
    rows.append(["成品转化累计", ""] + convert_row[1:])

    material_warnings = []
    if invalid_work_hours_rows:
        material_warnings.append(f"单日工时表中有 {invalid_work_hours_rows} 行无法识别，已沿用默认工时。")
    if ignored_work_hours_dates:
        preview = "、".join(f"{d.month}/{d.day}" for d in ignored_work_hours_dates[:5])
        material_warnings.append(f"有 {len(ignored_work_hours_dates)} 个工时日期不在排程周期内，已忽略：{preview}")
    if material_enabled and invalid_material_rows:
        material_warnings.append(f"物料交期表中有 {invalid_material_rows} 行无法识别，已忽略。")
    if material_enabled and ignored_material_dates:
        preview = "、".join(f"{d.month}/{d.day}" for d in ignored_material_dates[:5])
        material_warnings.append(f"有 {len(ignored_material_dates)} 个到料日期不在排程周期内，已忽略：{preview}")
    if material_enabled and material_limited_by_total and material_shortage_qty > 0:
        material_warnings.append(f"物料总量低于需求，缺少 {material_shortage_qty:,} 件物料；已按当前物料上限排产。")
    produced_total = sum(daily_total)
    if material_enabled and not any(not is_shift_empty(shift["daily_prod"]) for shift in shifts_production if shift["is_new"]):
        if produced_total >= production_target:
            min_gap = min(material_gap_row) if material_gap_row else None
            if min_gap is not None and min_gap > 0:
                run_mode = "场景一：物料充足，老班组月初连续正排，满足需求后释放"
                message = f"✅ 排产完成 | {run_mode}"
            else:
                run_mode = "场景三：物料间歇缺口，保留老班组放空并后续补产"
                message = f"✅ 排产完成 | {run_mode}"
    if produced_total < production_target:
        if material_enabled and material_limited_by_total:
            production_target = int(produced_total)
            message = f"✅ 排产完成 | 物料总量低于需求，已按可排物料上限输出"
        else:
            shortage = production_target - produced_total
            material_warnings.append(f"当前物料交期约束下仍有 {shortage:,} 件未排完，请补充到料、增加产能或延长周期。")
            message = f"⚠️ 排产未完全覆盖 | {run_mode} | 当前仍有{shortage:,}件未排完"

    schedule_df = pd.DataFrame(rows, columns=columns)
    return schedule_df, message, run_mode, production_target, total_workdays, default_single_shift_daily, total_exist_capacity, production_end_date, final_shift_total, material_warnings

# ------------------------------
# 四、页面配置
# ------------------------------
st.set_page_config(page_title="智能排产系统", page_icon="📊", layout="wide")
st.markdown(
    """
    <style>
    .material-input-panel {
        padding-top: 2px;
    }
    .material-input-panel div[data-testid="stHorizontalBlock"] {
        align-items: stretch;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(input[aria-label="启用物料交期约束"]) {
        background: #ffffff;
        border-color: #d8dee8;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(input[aria-label="启用物料交期约束"]) label,
    div[data-testid="stVerticalBlockBorderWrapper"]:has(input[aria-label="启用物料交期约束"]) p {
        color: #2f3340;
    }
    .material-rule-panel {
        display: flex;
        align-items: center;
        max-width: none;
        padding: 14px 16px;
        margin-bottom: 18px;
        border: 1px solid #dbe3ef;
        border-radius: 8px;
        background: #f7fbff;
        color: #1f3b57;
        line-height: 1.55;
        font-size: 13px;
        min-height: 100px;
    }
    .material-rule-panel ol {
        margin: 0;
        padding-left: 20px;
    }
    .material-rule-panel li {
        margin-bottom: 5px;
    }
    .material-rule-panel li:last-child {
        margin-bottom: 0;
    }
    .material-section-title {
        margin: 0 0 4px 0;
        font-weight: 800;
        color: #2f3340;
        line-height: 1.35;
    }
    .material-muted {
        color: #667085;
        font-size: 12px;
        margin-top: -1px;
        margin-bottom: 10px;
        line-height: 1.45;
    }
    .material-action-spacer {
        height: 14px;
    }
    .material-status-spacer {
        height: 16px;
    }
    div[data-testid="stButton"] > button {
        white-space: nowrap;
    }
    div[data-testid="stDownloadButton"] > button {
        width: 100%;
        min-height: 104px;
        height: 104px;
        border-color: #d9e2ef;
        background: #ffffff;
        color: #2f3340;
    }
    div[data-testid="stFileUploader"] {
        width: 100%;
    }
    div[data-testid="stFileUploader"] section {
        width: 100%;
        min-height: 104px;
        height: 104px;
        padding: 12px 150px 12px 16px;
        position: relative;
        overflow: hidden;
        border-color: #d9e2ef;
        background: #f3f6fb;
        color: #2f3340;
    }
    div[data-testid="stFileUploader"] section > div {
        width: 100%;
        display: flex;
        align-items: center;
        gap: 12px;
    }
    div[data-testid="stFileUploader"] section button {
        min-height: 40px;
        white-space: nowrap;
        position: absolute;
        right: 12px;
        top: 50%;
        transform: translateY(-50%);
        border: 1px solid #d0d5dd;
        background: #ffffff;
        color: #344054;
    }
    div[data-testid="stFileUploader"] small {
        color: #667085;
        line-height: 1.25;
    }
    .material-input-panel div[data-testid="stDataFrame"] {
        margin-top: 2px;
    }
    .hero-panel {
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 24px;
        padding: 24px 28px;
        margin: 2px 0 18px 0;
        border: 1px solid #e1e8f2;
        border-radius: 10px;
        background:
            linear-gradient(135deg, rgba(31, 115, 255, .08) 0%, rgba(255, 255, 255, 0) 40%),
            linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
        box-shadow: 0 12px 32px rgba(16, 24, 40, .055);
    }
    .hero-title {
        font-size: 30px;
        font-weight: 900;
        color: #101828;
        margin-bottom: 6px;
    }
    .hero-subtitle {
        color: #667085;
        font-size: 15px;
        line-height: 1.6;
    }
    .hero-meta {
        min-width: 180px;
        text-align: right;
        color: #175cd3;
        font-size: 13px;
        font-weight: 800;
        line-height: 1.6;
    }
    .overview-card {
        padding: 14px 16px;
        border: 1px solid #e1e8f2;
        border-radius: 8px;
        background: #ffffff;
        box-shadow: 0 10px 26px rgba(16, 24, 40, .04);
        min-height: 78px;
    }
    .overview-label {
        color: #667085;
        font-size: 13px;
        font-weight: 700;
    }
    .overview-value {
        color: #101828;
        font-size: 20px;
        font-weight: 900;
        margin-top: 6px;
        line-height: 1.25;
        white-space: nowrap;
    }
    .overview-value-date {
        font-size: 18px;
        letter-spacing: 0;
    }
    .form-section-kicker {
        color: #667085;
        font-size: 12px;
        font-weight: 800;
        letter-spacing: 0;
        margin-bottom: 2px;
    }
    .form-section-title {
        color: #101828;
        font-size: 18px;
        font-weight: 900;
        line-height: 1.35;
        margin-bottom: 4px;
    }
    .form-section-copy {
        color: #667085;
        font-size: 13px;
        line-height: 1.5;
        margin-bottom: 12px;
    }
    .capacity-summary-card {
        min-height: 136px;
        padding: 18px 20px;
        border: 1px solid #dbe7f5;
        border-radius: 8px;
        background: #f8fbff;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .capacity-summary-label {
        color: #667085;
        font-size: 13px;
        font-weight: 800;
        margin-bottom: 8px;
    }
    .capacity-summary-value {
        color: #101828;
        font-size: 34px;
        font-weight: 900;
        line-height: 1.1;
        letter-spacing: 0;
    }
    .capacity-summary-note {
        color: #667085;
        font-size: 12px;
        line-height: 1.45;
        margin-top: 10px;
    }
    .capacity-help {
        padding: 14px 16px;
        border: 1px solid #dbeafe;
        border-radius: 8px;
        background: #eef6ff;
        color: #1f3b57;
        line-height: 1.6;
        margin-top: 20px;
        margin-bottom: 12px;
    }
    .panel-bottom-spacer {
        height: 12px;
    }
    .capacity-upload-intro {
        padding: 14px 16px;
        border: 1px solid #dbeafe;
        border-radius: 8px;
        background: #f8fbff;
        color: #1f3b57;
        line-height: 1.55;
        margin: 14px 0 12px 0;
    }
    .capacity-upload-title {
        color: #2f3340;
        font-size: 14px;
        font-weight: 900;
        line-height: 1.35;
        margin-bottom: 4px;
    }
    .capacity-upload-copy {
        color: #667085;
        font-size: 12px;
        line-height: 1.45;
        margin-bottom: 8px;
    }
    .module-heading {
        display: flex;
        align-items: center;
        gap: 12px;
        margin: 24px 0 12px 0;
        padding: 12px 14px;
        border: 1px solid #e6edf7;
        border-radius: 8px;
        background: linear-gradient(90deg, #f8fbff 0%, #ffffff 72%);
    }
    .module-index {
        width: 32px;
        height: 32px;
        border-radius: 7px;
        background: #1f73ff;
        color: white;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-weight: 900;
    }
    .module-title {
        font-size: 18px;
        font-weight: 900;
        color: #101828;
    }
    .module-subtitle {
        color: #667085;
        font-size: 13px;
        margin-left: 6px;
    }
    div[data-testid="stMetric"] {
        padding: 10px 14px;
        border: 1px solid #e1e8f2;
        border-radius: 8px;
        background: white;
    }
    div[data-testid="stVerticalBlock"] > div:has(> div[data-testid="stSelectbox"]),
    div[data-testid="stVerticalBlock"] > div:has(> div[data-testid="stNumberInput"]),
    div[data-testid="stVerticalBlock"] > div:has(> div[data-testid="stDateInput"]) {
        margin-bottom: 4px;
    }
    div[data-testid="stButton"] > button,
    div[data-testid="stDownloadButton"] > button {
        border-radius: 8px;
        font-weight: 800;
    }
    div[data-testid="stButton"] > button[kind="primary"] {
        min-height: 48px;
        background: #1f73ff;
        border-color: #1f73ff;
        box-shadow: 0 10px 20px rgba(31, 115, 255, .18);
    }
    .block-container {
        padding-top: 22px;
        padding-bottom: 48px;
        max-width: 1500px;
    }
    #MainMenu, footer {
        visibility: hidden;
    }
    div[data-testid="stDataFrame"] {
        border-radius: 8px;
        overflow: hidden;
    }
    .result-intro {
        padding: 16px 18px;
        border: 1px solid #dbeafe;
        border-radius: 8px;
        background: #f8fbff;
        color: #1f3b57;
        line-height: 1.65;
        margin-bottom: 14px;
    }
    .section-spacer {
        height: 4px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    f"""
    <div class="hero-panel">
        <div>
            <div class="hero-title">智能排产系统</div>
            <div class="hero-subtitle">集中配置制程参数、工厂日历、单日工时、物料交期与爬坡规则，生成可执行的排产计划和结论建议。</div>
        </div>
        <div class="hero-meta">生产计划<br/>参数录入 · 计划测算 · 风险建议</div>
    </div>
    """,
    unsafe_allow_html=True,
)

overview_container = st.container()
main_col = st.container()

with main_col:
    st.markdown(
        "<div class='module-heading'><span class='module-index'>01</span><span class='module-title'>选择生产制程</span><span class='module-subtitle'>选择适用的生产制程模板</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        selected_process = st.selectbox(
            "模板",
            options=list(PROCESS_CONFIG.keys()),
            index=2,
            format_func=lambda x: f"{x} - {PROCESS_CONFIG[x]['desc']}"
        )
        current_process_config = PROCESS_CONFIG[selected_process]

    st.markdown(
        f"<div class='module-heading'><span class='module-index'>02</span><span class='module-title'>【{selected_process}】需求与周期输入</span><span class='module-subtitle'>配置需求、库存、班组与交付日期</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        target_col, date_col = st.columns(2, gap="large")
        with target_col:
            st.markdown("<div class='form-section-kicker'>DEMAND</div><div class='form-section-title'>需求与库存</div><div class='form-section-copy'>录入总需求、期初库存和特殊占用，系统据此计算最终生产目标。</div>", unsafe_allow_html=True)
            total_demand = st.number_input("总需求量", min_value=0, value=0, step=1000)
            initial_stock = st.number_input("半成品期初库存", min_value=0, value=0, step=1000)
            special_occupy = st.number_input("特殊占用（需生产）", min_value=0, value=0, step=100)
        with date_col:
            st.markdown("<div class='form-section-kicker'>PERIOD</div><div class='form-section-title'>日期与班组</div><div class='form-section-copy'>设置现有班组数量和排产周期，后续日历与产能按此范围展开。</div>", unsafe_allow_html=True)
            existing_shift_count = st.number_input("现有老班组数量", min_value=0, max_value=10, value=0)
            schedule_start_date = st.date_input("排产开始日期", value=None)
            demand_end_date = st.date_input("需求最终截止日期", value=None)
            date_period_ready = (
                schedule_start_date is not None
                and demand_end_date is not None
                and schedule_start_date <= demand_end_date
            )
            if schedule_start_date is None or demand_end_date is None:
                st.info("请先手动选择排产开始日期和需求最终截止日期；未选择前日期相关数值按 0 显示。")
            elif schedule_start_date > demand_end_date:
                st.error("排产开始日期不能晚于需求最终截止日期。")

    st.markdown(
        "<div class='module-heading'><span class='module-index'>03</span><span class='module-title'>工厂日历设置</span><span class='module-subtitle'>设置排产日历与休息日</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        default_sunday_rest = st.checkbox("默认周日自动设为休息日", value=False, disabled=not date_period_ready)
        full_calendar_dates = generate_full_date_list(schedule_start_date, demand_end_date) if date_period_ready else []
        if default_sunday_rest:
            default_rest_dates = [d for d in full_calendar_dates if d.weekday() == 6]
        else:
            default_rest_dates = []

        rest_dates = st.multiselect(
            "自定义休息日",
            options=full_calendar_dates,
            default=default_rest_dates,
            format_func=lambda x: f"{x.month}月{x.day}日 周{['一','二','三','四','五','六','日'][x.weekday()]}"
        )
        rest_dates_set = set(rest_dates)
        workday_count = sum(1 for d in full_calendar_dates if d not in rest_dates_set)
        if date_period_ready:
            st.info(f"排产范围：{schedule_start_date.strftime('%Y-%m-%d')} 至 {demand_end_date.strftime('%Y-%m-%d')}，共 {len(full_calendar_dates)} 天，工作日 {workday_count} 天，休息日 {len(rest_dates_set)} 天")
        else:
            st.info("排产范围：待选择，共 0 天，工作日 0 天，休息日 0 天")

    st.markdown(
        "<div class='module-heading'><span class='module-index'>04</span><span class='module-title'>制程能力输入</span><span class='module-subtitle'>合并制程能力与单日工时，实时核算可用产能</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        capability_col, hours_col, capacity_col = st.columns([1, 1, 1], gap="large")
        with capability_col:
            st.markdown("<div class='form-section-kicker'>PROCESS</div><div class='form-section-title'>基础能力</div><div class='form-section-copy'>设置制程转换提前期和单班组小时产出。</div>", unsafe_allow_html=True)
            lead_time_days = st.number_input("成品转化Lead Time(工作日)", min_value=0, value=0)
            uph_base = st.number_input("单班组UPH", min_value=0, value=0)
            idle_convert_threshold_days = st.number_input("放空转新阈值(工作日)", min_value=0, max_value=30, value=5, step=1)
        with hours_col:
            st.markdown("<div class='form-section-kicker'>HOURS</div><div class='form-section-title'>工时策略</div><div class='form-section-copy'>统一工时适合稳定节奏，按日期编辑适合爬坡或临时调整。</div>", unsafe_allow_html=True)
            work_hours_mode = st.radio(
                "工时输入方式",
                options=["统一工时", "按日期编辑"],
                horizontal=True,
            )
            work_hours = st.number_input(
                "默认单班组单日工时",
                min_value=0,
                max_value=24,
                value=0,
                step=1,
            )
        with capacity_col:
            default_shift_capacity = uph_base * work_hours
            st.markdown(
                f"""
                <div class="capacity-summary-card">
                    <div class="capacity-summary-label">默认单班组单日满产</div>
                    <div class="capacity-summary-value">{default_shift_capacity:,}</div>
                    <div class="capacity-summary-note">计算公式：单班组 UPH {uph_base:,} × 默认单日工时 {work_hours} 小时</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        work_hours_plan_df = pd.DataFrame({
            "日期": full_calendar_dates,
            "是否休息日": ["是" if d in rest_dates_set else "否" for d in full_calendar_dates],
            "单班组单日工时": [0 if d in rest_dates_set else int(work_hours) for d in full_calendar_dates],
        })
        if work_hours_mode == "按日期编辑":
            if date_period_ready:
                work_hours_template = build_work_hours_template(
                    schedule_start_date,
                    demand_end_date,
                    work_hours,
                )
                st.markdown(
                    "<div class='capacity-upload-intro'>按日期编辑时，请先下载横向工时模板，在每个日期下方填写单班组单日工时，再上传文件参与产能核算。</div>",
                    unsafe_allow_html=True,
                )
                hours_template_col, hours_upload_col, hours_preview_col = st.columns([0.26, 0.36, 0.38], gap="medium")
                with hours_template_col:
                    st.markdown("<div class='capacity-upload-title'>横向模板</div><div class='capacity-upload-copy'>日期横向展开，适合批量维护每日工时。</div>", unsafe_allow_html=True)
                    st.download_button(
                        "下载工时模板",
                        data=work_hours_template,
                        file_name=f"单日工时输入模板_{schedule_start_date.strftime('%Y%m%d')}_{demand_end_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with hours_upload_col:
                    st.markdown("<div class='capacity-upload-title'>上传工时</div><div class='capacity-upload-copy'>上传后覆盖默认工时，空白单元格沿用默认值。</div>", unsafe_allow_html=True)
                    uploaded_work_hours_file = st.file_uploader(
                        "上传已填写的单日工时Excel",
                        type=["xlsx", "xls"],
                        key="work_hours_upload",
                        label_visibility="collapsed",
                    )
                if uploaded_work_hours_file is not None:
                    try:
                        uploaded_work_hours_df, work_hours_messages = parse_work_hours_upload(
                            uploaded_work_hours_file,
                            schedule_start_date,
                            demand_end_date,
                            work_hours,
                        )
                        work_hours_plan_df = uploaded_work_hours_df.copy()
                        for msg in work_hours_messages:
                            st.warning(msg)
                        st.success(f"已读取单日工时：{len(work_hours_plan_df)} 条。")
                    except Exception as exc:
                        st.error(f"单日工时Excel解析失败：{exc}")
                        st.stop()
                else:
                    work_hours_plan_df = work_hours_plan_df[["日期", "单班组单日工时"]].copy()
                with hours_preview_col:
                    st.markdown("<div class='capacity-upload-title'>工时预览</div><div class='capacity-upload-copy'>展示当前周期内每日单班组工时。</div>", unsafe_allow_html=True)
                    st.dataframe(
                        work_hours_plan_df,
                        use_container_width=True,
                        hide_index=True,
                        height=150,
                        column_config={
                            "日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD"),
                            "单班组单日工时": st.column_config.NumberColumn("单班组单日工时", min_value=0, max_value=24, step=1),
                        },
                    )
            else:
                st.info("请先手动选择排产开始日期和需求最终截止日期；日期未完整选择前，工时明细为 0 条。")
                work_hours_plan_df = work_hours_plan_df[["日期", "单班组单日工时"]].copy()
        else:
            st.markdown(
                "<div class='capacity-help'>当前按统一工时计算；如前期 8 小时、后期满产，可切换为“按日期编辑”直接修改每天工时。</div>",
                unsafe_allow_html=True,
            )
            work_hours_plan_df = work_hours_plan_df[["日期", "单班组单日工时"]].copy()
        st.markdown("<div class='panel-bottom-spacer'></div>", unsafe_allow_html=True)

    st.markdown(
        "<div class='module-heading'><span class='module-index'>05</span><span class='module-title'>物料交期输入</span><span class='module-subtitle'>导入物料到料节奏，启用后参与排产约束</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        material_enabled = st.checkbox("启用物料交期约束", value=True)
        default_material_initial_stock = 0
        material_initial_stock = 0
        material_plan_df = pd.DataFrame({
            "日期": full_calendar_dates,
            "预计到料数量": [0] * len(full_calendar_dates),
        })
        material_schedule_enabled = False
        if material_enabled:
            default_material_initial_stock = 0
            material_initial_stock = int(default_material_initial_stock)
            if "material_upload_nonce" not in st.session_state:
                st.session_state.material_upload_nonce = 0

            if date_period_ready:
                material_template = build_material_template(
                    schedule_start_date,
                    demand_end_date,
                    default_material_initial_stock,
                )
                st.markdown("<div class='material-input-panel'>", unsafe_allow_html=True)
                action_col, preview_col = st.columns([0.44, 0.56], gap="large")
                with action_col:
                    st.markdown("<div class='material-section-title'>Excel模板与上传</div>", unsafe_allow_html=True)
                    st.markdown("<div class='material-muted'>先下载模板填写物料期初库存和每日到料数量，再上传文件参与排产约束</div>", unsafe_allow_html=True)
                    download_inner_col, upload_inner_col = st.columns([0.36, 0.64], gap="small")
                    with download_inner_col:
                        st.download_button(
                            "下载模板",
                            data=material_template,
                            file_name=f"物料交期输入模板_{schedule_start_date.strftime('%Y%m%d')}_{demand_end_date.strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                    with upload_inner_col:
                        uploaded_material_file = st.file_uploader(
                            "上传已填写的物料交期Excel",
                            type=["xlsx", "xls"],
                            key=f"material_upload_{st.session_state.material_upload_nonce}",
                            label_visibility="collapsed",
                        )
                        if uploaded_material_file is not None:
                            delete_uploaded_material = st.button(
                                "删除文件",
                                key=f"delete_material_upload_{st.session_state.material_upload_nonce}",
                                use_container_width=True,
                            )
                            if delete_uploaded_material:
                                st.session_state.material_upload_nonce += 1
                                st.rerun()
                    st.markdown("<div class='material-action-spacer'></div>", unsafe_allow_html=True)
                    st.markdown("<div class='material-section-title'>到料规则</div>", unsafe_allow_html=True)
                    st.markdown("<div class='material-muted'>排产前按工厂工作日偏移核算可用物料</div>", unsafe_allow_html=True)
                    st.markdown(
                        """
                        <div class="material-rule-panel">
                            <ol>
                                <li><strong>到料下个工作日可用：</strong>T 日到厂物料，最早从第 03 步工厂日历里的下一个工作日投入排产。</li>
                                <li><strong>排产前核算：</strong>当日可用物料 = 上一个排产日结余物料 + 当日已生效到料数量。</li>
                            </ol>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                if uploaded_material_file is not None:
                    st.markdown("<div class='material-status-spacer'></div>", unsafe_allow_html=True)
                    try:
                        material_initial_stock, material_plan_df, upload_messages = parse_material_upload(
                            uploaded_material_file,
                            schedule_start_date,
                            demand_end_date,
                            default_material_initial_stock,
                        )
                        material_schedule_enabled = True
                        st.success(f"已读取物料期初库存：{material_initial_stock:,}，到料日期 {len(material_plan_df)} 条。")
                        for msg in upload_messages:
                            st.warning(msg)
                    except Exception as exc:
                        st.error(f"物料交期Excel解析失败：{exc}")
                        st.stop()
                with preview_col:
                    st.markdown("<div class='material-section-title'>物料交期预览</div>", unsafe_allow_html=True)
                    st.markdown("<div class='material-muted'>展示当前排产周期内每日预计到料数量</div>", unsafe_allow_html=True)
                    st.dataframe(
                        material_plan_df,
                        use_container_width=True,
                        hide_index=True,
                        height=285,
                        column_config={
                            "日期": st.column_config.DateColumn("日期", format="YYYY-MM-DD"),
                            "预计到料数量": st.column_config.NumberColumn("预计到料数量", min_value=0, step=1000),
                        },
                    )
                st.markdown("</div>", unsafe_allow_html=True)
                st.markdown("<div class='panel-bottom-spacer'></div>", unsafe_allow_html=True)
            else:
                st.info("请先手动选择排产开始日期和需求最终截止日期；未选择前物料交期明细为 0 条。")
        else:
            st.info("已关闭物料交期约束，排产不受物料到料限制。")

    st.markdown(
        "<div class='module-heading'><span class='module-index'>06</span><span class='module-title'>爬坡规则配置</span><span class='module-subtitle'>仅新增班组适用</span></div>",
        unsafe_allow_html=True,
    )
    with st.container(border=True):
        col1, col2 = st.columns([.42, 1.58], gap="large")
        with col1:
            default_model = list(RAMP_DATA.keys())[0]
            selected_model = st.selectbox("选择机型", options=list(RAMP_DATA.keys()), index=list(RAMP_DATA.keys()).index(default_model))
            default_ratio = list(RAMP_DATA[selected_model].keys())[0]
            new_human_ratio = st.selectbox("选择新人占比", options=list(RAMP_DATA[selected_model].keys()), index=list(RAMP_DATA[selected_model].keys()).index(default_ratio))
        with col2:
            current_ramp_curve = RAMP_DATA[selected_model][new_human_ratio]
            ramp_df = pd.DataFrame({
                "爬坡天数": [f"DAY{i+1}" for i in range(len(current_ramp_curve))],
                "产能比例": [f"{x}%" for x in current_ramp_curve]
            }).T
            st.dataframe(ramp_df, use_container_width=True, hide_index=True, height=126)
            st.caption(f"爬坡总天数：{len(current_ramp_curve)}天")

with overview_container:
    metric_cols = st.columns([1.35, 1, 1, 1, 1], gap="medium")
    period_value = (
        f"{schedule_start_date.strftime('%Y/%m/%d')} 至 {demand_end_date.strftime('%m/%d')}"
        if date_period_ready
        else "待选择"
    )
    metric_values = [
        ("排产范围", period_value, "overview-value overview-value-date"),
        ("排产天数", f"{workday_count} 个排产日"),
        ("单日工时", f"{work_hours} 小时"),
        ("单班组满产", f"{uph_base * work_hours:,}"),
        ("单班组UPH", f"{uph_base:,}"),
    ]
    for metric_col, metric_item in zip(metric_cols, metric_values):
        if len(metric_item) == 3:
            label, value, value_class = metric_item
        else:
            label, value = metric_item
            value_class = "overview-value"
        with metric_col:
            st.markdown(f"<div class='overview-card'><div class='overview-label'>{label}</div><div class='{value_class}'>{value}</div></div>", unsafe_allow_html=True)
    st.markdown("<div style='height: 14px'></div>", unsafe_allow_html=True)

st.markdown(
    "<div class='module-heading'><span class='module-index'>07</span><span class='module-title'>排产结果</span><span class='module-subtitle'>生成计划、结论建议并导出 Excel</span></div>",
    unsafe_allow_html=True,
)
st.markdown(
    """
    <div class="result-intro">
        完成上方配置后点击开始排产。系统会同步输出运行模式、关键测算结果、结论建议、排产明细表，并支持导出 Excel。
    </div>
    """,
    unsafe_allow_html=True,
)
if not date_period_ready:
    st.info("请先手动选择完整且有效的排产日期范围，再开始排产。")

if st.button(f"开始【{selected_process}】制程排产", type="primary", use_container_width=True, disabled=not date_period_ready):
    schedule_df, message, mode, production_target, total_workdays, single_shift_daily, total_exist_capacity, production_end_date, final_shift_total, material_warnings = schedule_engine(
        process_name=selected_process,
        total_demand=total_demand,
        initial_stock=initial_stock,
        special_occupy=special_occupy,
        uph_base=uph_base,
        work_hours=work_hours,
        existing_shift_count=existing_shift_count,
        schedule_start_date=schedule_start_date,
        demand_end_date=demand_end_date,
        lead_time_days=lead_time_days,
        rest_dates_set=rest_dates_set,
        material_initial_stock=material_initial_stock,
        material_plan_df=material_plan_df,
        selected_model=selected_model,
        new_human_ratio=new_human_ratio,
        material_enabled=material_schedule_enabled,
        work_hours_plan_df=work_hours_plan_df,
        idle_convert_threshold_days=idle_convert_threshold_days,
    )

    # 加班 / 减班优化提示
    overtime_candidate_qty, overtime_available_shift_count, overtime_candidate_label = find_overtime_reduction_candidate(schedule_df)
    can_optimize, optimize_suggest, selected_dates = analyze_overtime_optimization(
        overtime_candidate_qty,
        overtime_available_shift_count,
        single_shift_daily,
        production_end_date,
        rest_dates_set,
        full_calendar_dates,
        overtime_candidate_label,
    )
    if can_optimize:
        st.warning(optimize_suggest)

    # 核心计算结果
    st.markdown("### 核心计算结果")
    st.markdown(f"- **总需求量**：{total_demand:,}")
    st.markdown(f"- **半成品期初库存**：{initial_stock:,}")
    if material_schedule_enabled:
        st.markdown(f"- **物料期初库存**：{material_initial_stock:,}")
    else:
        st.markdown("- **物料交期约束**：未启用")
    st.markdown(f"- **特殊占用（需生产）**：{special_occupy:,}")
    st.markdown(f"- **最终生产目标**：{production_target:,}")
    st.markdown(f"- **现有班组总产能**：{total_exist_capacity:,}")
    st.markdown(f"- **周期有效工作日**：{total_workdays}天")
    st.markdown(f"- **需求最终截止日**：{demand_end_date.month}月{demand_end_date.day}日")
    st.markdown(f"- **生产必须完成截止日**：{production_end_date.month}月{production_end_date.day}日（提前{lead_time_days}个工作日，休息日不计入）")
    st.divider()

    if mode == "small_gap":
        st.warning("排产未完全覆盖")
    elif "错误" in message:
        st.error("排产错误")
    elif "未完全覆盖" in message:
        st.warning("排产未完全覆盖")
    else:
        st.success("排产完成")
    if "错误" not in message:
        insight_summary, conclusions, suggestions = build_schedule_insights(
            schedule_df=schedule_df,
            production_target=production_target,
            total_demand=total_demand,
            initial_stock=initial_stock,
            special_occupy=special_occupy,
            total_exist_capacity=total_exist_capacity,
            existing_shift_count=existing_shift_count,
            uph_base=uph_base,
            material_schedule_enabled=material_schedule_enabled,
            material_warnings=material_warnings,
            mode=mode,
            idle_convert_threshold_days=idle_convert_threshold_days,
        )
        if insight_summary:
            st.markdown("### 排产结论与建议")
            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
            metric_col1.metric("计划产出", f"{insight_summary['planned_output']:,}")
            if insight_summary.get("material_shortage_qty", 0) > 0:
                metric_col2.metric("需求差额", f"缺少 {insight_summary['material_shortage_qty']:,}")
            elif insight_summary["shortage"] > 0:
                metric_col2.metric("目标差异", f"缺口 {insight_summary['shortage']:,}")
            else:
                metric_col2.metric("目标差异", "已达成")
            metric_col3.metric(
                "启用班组",
                f"老{insight_summary['old_shift_count']} / 新{insight_summary['new_shift_count']}",
            )
            metric_col4.metric("最后生产日", insight_summary["last_prod_day"])

            if material_schedule_enabled and insight_summary["material_gap_min"] is not None:
                st.caption(f"最低物料 gap：{insight_summary['material_gap_min']:,}")

            conclusion_col, suggestion_col = st.columns(2)
            with conclusion_col:
                st.markdown("**结论**")
                for item in conclusions:
                    st.markdown(f"- {item}")
            with suggestion_col:
                st.markdown("**建议**")
                for item in suggestions:
                    st.markdown(f"- {item}")
            st.divider()

        st.dataframe(schedule_df, use_container_width=True, height=500)

        # Excel导出
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            schedule_df.to_excel(writer, sheet_name=f'{selected_process}排产表', index=False)
            work_hours_plan_df.to_excel(writer, sheet_name='单日工时输入', index=False)
            if material_schedule_enabled:
                material_plan_df.to_excel(writer, sheet_name='物料交期输入', index=False)
        st.download_button(
            label=f"下载【{selected_process}】制程排产表Excel",
            data=buffer.getvalue(),
            file_name=f"{selected_process}制程_智能排产计划表_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        for msg in material_warnings:
            st.warning(msg)
